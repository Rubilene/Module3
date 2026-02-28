"""
src/07_dashboard_report.py
===========================
Automated KPI Report Generator

Queries the analytical warehouse and produces:
  1. An Excel workbook with formatted KPI tables (kpi_report.xlsx)
  2. Matplotlib chart images embedded in the report
  3. Console summary output

In production this would integrate with Power BI via REST API
or generate scheduled PDF reports for distribution.
"""

import os
import sys
import sqlite3
import logging
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime

warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.settings import WAREHOUSE_DB, REPORT_FILE, REPORTS_DIR

log = logging.getLogger("etl.report")

NAVY  = "1F3864"
STEEL = "2E5D9E"
TEAL  = "006B75"
WHITE = "FFFFFF"
LGRAY = "F2F4F7"

# ── Warehouse query helpers ───────────────────────────────────────
def query(conn, sql: str) -> pd.DataFrame:
    return pd.read_sql(sql, conn)

def load_kpis(conn: sqlite3.Connection) -> dict:
    """Run all KPI queries against the warehouse."""

    kpis = {}

    kpis["summary"] = query(conn, """
        SELECT
            COUNT(*)                                    AS total_studies,
            ROUND(AVG(turnaround_mins), 1)              AS avg_tat_mins,
            ROUND(AVG(turnaround_mins)/60.0, 2)         AS avg_tat_hours,
            SUM(report_issued_flag)                     AS reports_issued,
            ROUND(AVG(report_issued_flag)*100, 1)       AS report_rate_pct,
            COUNT(DISTINCT radiologist_key)             AS active_radiologists
        FROM Fact_Radiology
        WHERE tat_flag = 'OK'
    """)

    kpis["by_modality"] = query(conn, """
        SELECT
            m.modality_name,
            m.modality_group,
            COUNT(*)                               AS study_count,
            ROUND(AVG(f.turnaround_mins), 1)       AS avg_tat_mins,
            SUM(f.report_issued_flag)              AS reports_issued,
            ROUND(AVG(f.report_issued_flag)*100,1) AS report_rate_pct
        FROM Fact_Radiology f
        JOIN Dim_Modality m ON f.modality_key = m.modality_key
        WHERE f.tat_flag = 'OK'
        GROUP BY m.modality_name, m.modality_group
        ORDER BY study_count DESC
    """)

    kpis["by_site"] = query(conn, """
        SELECT
            l.site_code,
            l.department,
            COUNT(*)                               AS study_count,
            ROUND(AVG(f.turnaround_mins), 1)       AS avg_tat_mins
        FROM Fact_Radiology f
        JOIN Dim_Location l ON f.location_key = l.location_key
        WHERE f.tat_flag = 'OK'
        GROUP BY l.site_code, l.department
        ORDER BY study_count DESC
    """)

    kpis["tat_distribution"] = query(conn, """
        SELECT
            tat_band,
            COUNT(*) AS study_count
        FROM Fact_Radiology
        WHERE tat_flag = 'OK' AND tat_band != 'Invalid'
        GROUP BY tat_band
        ORDER BY
            CASE tat_band
                WHEN '< 1 hour'  THEN 1
                WHEN '1–4 hours' THEN 2
                WHEN '4–24 hours' THEN 3
                WHEN '1–3 days'  THEN 4
                WHEN '> 3 days'  THEN 5
                ELSE 6
            END
    """)

    kpis["by_priority"] = query(conn, """
        SELECT
            priority,
            COUNT(*)                         AS study_count,
            ROUND(AVG(turnaround_mins), 1)   AS avg_tat_mins
        FROM Fact_Radiology
        WHERE tat_flag = 'OK'
        GROUP BY priority
        ORDER BY study_count DESC
    """)

    kpis["monthly_trend"] = query(conn, """
        SELECT
            SUBSTR(CAST(date_key AS TEXT), 1, 6) AS year_month,
            COUNT(*)                              AS study_count,
            ROUND(AVG(turnaround_mins), 1)        AS avg_tat_mins
        FROM Fact_Radiology
        WHERE tat_flag = 'OK' AND date_key IS NOT NULL
        GROUP BY year_month
        ORDER BY year_month
    """)

    kpis["data_quality"] = query(conn, """
        SELECT
            tat_flag,
            COUNT(*) AS record_count
        FROM Fact_Radiology
        GROUP BY tat_flag
    """)

    return kpis

# ── Excel report builder ──────────────────────────────────────────
def write_excel_report(kpis: dict, filepath: str):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Helpers ───────────────────────────────────────────────────
    def header_fill(color=NAVY):
        return PatternFill("solid", fgColor=color)

    def alt_fill(i, color1=WHITE, color2="EEF2F7"):
        return PatternFill("solid", fgColor=color1 if i % 2 == 0 else color2)

    def border():
        thin = Side(style="thin", color="BFCFE8")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_table(ws, df, start_row=3, title=None):
        if title:
            ws.cell(row=start_row-1, column=1, value=title).font = Font(
                bold=True, size=12, color=NAVY)
        # Headers
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=ci, value=col.replace("_", " ").title())
            cell.fill      = header_fill()
            cell.font      = Font(bold=True, color=WHITE, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border()
        # Data rows
        for ri, row in enumerate(df.itertuples(index=False), start_row+1):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill      = alt_fill(ri)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border()
                if isinstance(val, float):
                    cell.number_format = "0.00"
        # Column widths
        for ci in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(ci)].width = 18
        return start_row + len(df) + 1

    # ── Sheet 1: Executive Summary ────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 20

    # Title banner
    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value      = "Radiology ETL Platform – KPI Dashboard Report"
    title_cell.fill       = header_fill(NAVY)
    title_cell.font       = Font(bold=True, color=WHITE, size=14)
    title_cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:F2")
    sub = ws1["A2"]
    sub.value     = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}  |  Multiverse ADF – Data Engineering PoC"
    sub.fill      = header_fill(STEEL)
    sub.font      = Font(color=WHITE, size=10, italic=True)
    sub.alignment = Alignment(horizontal="center")

    # KPI metric cards
    s = kpis["summary"].iloc[0]
    kpi_cards = [
        ("Total Studies Loaded",     int(s["total_studies"]),         ""),
        ("Average TAT (hours)",      float(s["avg_tat_hours"]),        "hrs"),
        ("Reports Issued",           int(s["reports_issued"]),         ""),
        ("Report Completion Rate",   float(s["report_rate_pct"]),      "%"),
        ("Active Radiologists",      int(s["active_radiologists"]),    ""),
    ]
    ws1.cell(row=4, column=1, value="Key Performance Indicators").font = \
        Font(bold=True, size=12, color=NAVY)
    for i, (lbl, val, unit) in enumerate(kpi_cards, 5):
        ws1.cell(row=i, column=1, value=lbl).font = Font(bold=True, color=NAVY, size=11)
        v = ws1.cell(row=i, column=2,
                     value=f"{val:,.1f}{unit}" if isinstance(val, float) else f"{val:,}{unit}")
        v.font      = Font(bold=True, size=11, color=STEEL)
        v.fill      = PatternFill("solid", fgColor="EEF2F7")
        v.alignment = Alignment(horizontal="center")
        v.border    = border()
        ws1.row_dimensions[i].height = 20

    # ── Sheet 2: Modality Analysis ────────────────────────────────
    ws2 = wb.create_sheet("Modality Analysis")
    ws2.sheet_view.showGridLines = False
    write_table(ws2, kpis["by_modality"], start_row=3, title="Study Volume & TAT by Modality")

    # ── Sheet 3: Site Performance ─────────────────────────────────
    ws3 = wb.create_sheet("Site Performance")
    ws3.sheet_view.showGridLines = False
    write_table(ws3, kpis["by_site"], start_row=3, title="Study Volume & TAT by Site")

    # ── Sheet 4: TAT Distribution ─────────────────────────────────
    ws4 = wb.create_sheet("TAT Distribution")
    ws4.sheet_view.showGridLines = False
    write_table(ws4, kpis["tat_distribution"], start_row=3, title="Turnaround Time Distribution")

    # ── Sheet 5: Monthly Trend ────────────────────────────────────
    ws5 = wb.create_sheet("Monthly Trend")
    ws5.sheet_view.showGridLines = False
    write_table(ws5, kpis["monthly_trend"], start_row=3, title="Monthly Study Volume & TAT Trend")

    # ── Sheet 6: Priority Analysis ────────────────────────────────
    ws6 = wb.create_sheet("Priority Analysis")
    ws6.sheet_view.showGridLines = False
    write_table(ws6, kpis["by_priority"], start_row=3, title="Study Volume & TAT by Priority")

    # ── Sheet 7: Data Quality ─────────────────────────────────────
    ws7 = wb.create_sheet("Data Quality")
    ws7.sheet_view.showGridLines = False
    write_table(ws7, kpis["data_quality"], start_row=3, title="Validation Flag Distribution")

    wb.save(filepath)
    log.info(f"  Excel report saved → {os.path.basename(filepath)}")

# ── Console KPI output ────────────────────────────────────────────
def print_kpi_console(kpis: dict):
    try:
        from tabulate import tabulate
        s = kpis["summary"].iloc[0]
        log.info("  ── KPI Summary ─────────────────────────────────")
        log.info(f"  Total studies in warehouse : {int(s['total_studies']):,}")
        log.info(f"  Average TAT                : {float(s['avg_tat_hours']):.2f} hours")
        log.info(f"  Reports issued             : {int(s['reports_issued']):,}")
        log.info(f"  Report completion rate     : {float(s['report_rate_pct'])}%")
        log.info(f"  Active radiologists        : {int(s['active_radiologists'])}")
        log.info("\n  ── Studies by Modality ─────────────────────────")
        tbl = kpis["by_modality"][["modality_name", "study_count", "avg_tat_mins"]].head(8)
        log.info("\n" + tabulate(tbl, headers="keys", tablefmt="simple", showindex=False))
        log.info("\n  ── TAT Distribution ────────────────────────────")
        tbl2 = kpis["tat_distribution"]
        log.info("\n" + tabulate(tbl2, headers="keys", tablefmt="simple", showindex=False))
    except ImportError:
        log.info(kpis["summary"].to_string())

# ── Main ──────────────────────────────────────────────────────────
def run(context=None):
    log.info("=" * 60)
    log.info("STEP 7 – KPI Dashboard Report Generation")
    log.info("=" * 60)

    conn  = sqlite3.connect(WAREHOUSE_DB)
    kpis  = load_kpis(conn)
    conn.close()

    print_kpi_console(kpis)
    write_excel_report(kpis, REPORT_FILE)

    log.info(f"  Report generation complete → {REPORT_FILE}\n")
    return kpis

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(message)s")
    run()
